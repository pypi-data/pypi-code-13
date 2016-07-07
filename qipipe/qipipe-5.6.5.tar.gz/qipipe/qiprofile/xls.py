import re
import openpyxl
from contextlib import contextmanager
from datetime import datetime
from bunch import bunchify
import inflection
from qiutil.logging import logger
from . import parse


class XLSError(Exception):
    pass


def load_workbook(filename):
    """
    :param filename: the XLS workbook file location
    :return the read-only ``openpyxl`` workbook object
    """
    return openpyxl.load_workbook(filename, read_only=True)


class Worksheet(object):
    
    PAREN_SUFFIX_REGEX = re.compile('\s*\(.*\)$')
    """Matches a parenthesis suffix."""
    
    NONWORD_REGEX = re.compile('\W+')
    """Matches a non-word string."""
    
    """
    Excel worksheet facade. The worksheet rows are required to consist of
    a column heading row followed by the data rows.
    """
    
    def __init__(self, workbook, sheet, *classes, **opts):
        """
        :param workbook: the openpyxl workbook
        :param sheet: the sheet name
        :param classes: the classes which have spreadsheet attributes
        :param opts: the following keyword arguments:
        :option parsers: the non-standard
            {attribute: value parser function} parsers dictionary
        :param column_attributes: the non-standard
            {column name: attribute} dictionary
        :raise XLSError: if there are unparsed columns
        """
        self.worksheet = workbook[sheet]
        """The wrapped openpyxl worksheet."""
        
        # The default parsers.
        parsers = parse.default_parsers(*classes)
        # Every worksheet has a subject number field.
        parsers.update(subject_number=int)
        # Add the special parsers.
        parsers_opt = opts.get('parsers', None)
        if parsers_opt:
            parsers.update(parsers_opt)
        self._parsers = parsers
        """The {attribute: value parser function} dictionary,"""
        
        # The first row is the headings.
        hdg_row = next(self.worksheet.iter_rows())
        # The column headings stop when the first blank cell is reached.
        headings = []
        for cell in hdg_row:
            if not cell.value:
                break
            headings.append(cell.value)
        # The non-standard {column name: attribute} dictionary.
        col_attrs = opts.get('column_attributes', {})
        # The attributes in the same order as the headings. The default
        # column attribute is generated by the _attributize function.
        self.attributes = [col_attrs.get(hdg, self._attributize(hdg))
                           for hdg in headings]
        """The row attributes in column order."""
        
        # There must be a parser for each attribute.
        unparsed = [attr for attr in self.attributes
                    if not attr in self._parsers]
        if unparsed:
            raise XLSError("%s columns are missing a parser: %s" %
                           (sheet, unparsed))
    
    def read(self, **condition):
        """
        Returns a row generator, where each row is a {attribute: value}
        bunch. This generator yields each row which satisfies the
        following condition:
        1. the row is non-empty, i.e. has at least one cell value, and
        2. if this reader has a filter, then the row satisfies the
           filter condition
        
        :param condition: the {attribute: value} row filter condition
        :return: the row {attribute: value} bunch generator
        """
        reader = Reader(self.worksheet, self.attributes, **condition)
        
        return (self._parse_row(row) for row in reader.read())
    
    def _parse_row(self, row):
        """
        Extracts and parses the row values.
        
        :param row: the XLS row to parse
        :return: the row {attribute: value} bunch
        """
        # The {attribute: value} dictionary.
        attr_val_dict = {}
        for i, cell in enumerate(row):
            # Stop when we are past the attributes.
            if i >= len(self.attributes):
                break
            # The attribute key.
            attr = self.attributes[i]
            # The value parser.
            parser = self._parsers[attr]
            # The attribute value.
            if cell.value == None or cell.value == '':
                val = None
            else:
                val = parser(cell.value)
            # Add the attribute value entry.
            attr_val_dict[attr] = val
        
        # Return the {attribute: value} bunch.
        return bunchify(attr_val_dict)
    
    def _attributize(self, s):
        """
        Converts the given string to an underscore attribute.
        A trailing parenthesized word is not included.
        
        Examples::
            
            >> reader._attributize('Start Date')
            start_date
            >> reader._attributize('Amount (mg/kg)')
            amount
        
        :param s: the string to convert
        :return: the underscore attribute name
        """
        stripped = Worksheet.PAREN_SUFFIX_REGEX.sub('', s)
        word = Worksheet.NONWORD_REGEX.sub('_', stripped)
        
        return inflection.underscore(word)


class Reader(object):
    """Reads an Excel worksheet."""
    
    def __init__(self, worksheet, attributes, **condition):
        """
        :param worksheet: the :attr:`worksheet` object
        :param conditional: the optional {attribute: value} row filter
            condition
        """
        self.worksheet = worksheet
        """The wrapped openpyxl worksheet."""
        
        # Validate the condition.
        for key in condition:
            if not key in attributes:
                raise XLSError("The condition key %s is not in the attribute"
                                " list" % key)
        # Capture the attribute range for iteration reuse.
        self._attr_range = range(0, len(attributes))
        # Convert the {attribute: value} dictionary to a
        # {column index: value} dictonary.
        self._filter = {attributes.index(attr): val
                        for attr, val in condition.iteritems()}
    
    def read(self):
        """
        Returns a row generator, where each row is a {attribute: value}
        bunch. This generator yields each row which satisfies the
        following condition:
        1. the row is non-empty, i.e. has at least one cell value, and
        2. if this reader has a filter, then the row satisfies the
           filter condition
        
        :return: the filtered ``openpyxl`` row iterator
        """
        # Start reading after the header row.
        row_iter = self.worksheet.iter_rows(row_offset=1)
        return (row for row in row_iter if self._include_row(row))
    
    def _include_row(self, row):
        """
        :param row: the row to check
        :return: whether to include the row in the iteration
        """
        return self._is_row_nonempty(row) and self._filter_row(row)
    
    def _is_row_nonempty(self, row):
        """
        :param row: the row to check
        :return: whether the row has at least one cell value
        """
        return any(row[i].value != None for i in self._attr_range)
    
    def _filter_row(self, row):
        """
        :param row: the row to check
        :return: whether the row satisfies the filter condition
        """
        return all(row[i].value == val for i, val in self._filter.iteritems())        
